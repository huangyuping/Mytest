"""
-------------------------------------------------
   File Name：     Do_Excel
   Description :
   Author :       小鱼
   date：          2018/3/19
-------------------------------------------------
   Change Activity:
                   2018/3/19:
-------------------------------------------------
"""
__author__ = '小鱼'

from openpyxl import load_workbook

class DoExcel:

    def __init__(self,excelpath):
        self.wb = load_workbook(excelpath)
        self.excelpath = excelpath
        self.sh_id = self.wb.get_sheet_by_name("id")
        self.sh_move = self.wb.get_sheet_by_name("move")

    def get_id(self):
        l=[]
        for i in range(1,self.sh_id.max_row+1):
            id = self.sh_id.cell(row=i,column=1).value
            l.append(id)
        return l