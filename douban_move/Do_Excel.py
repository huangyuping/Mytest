"""
-------------------------------------------------
   File Name：     Do_Excel
   Description :
   Author :       zws
   date：          2018/3/19
-------------------------------------------------
   Change Activity:
                   2018/3/19:
-------------------------------------------------
"""
__author__ = 'zws'

from openpyxl import load_workbook

class DoExcel:

    def __init__(self,excelpath):
        self.wb = load_workbook(excelpath)
        self.excelpath = excelpath
        self.sh_id = self.wb.get_sheet_by_name("id")
        self.sh_move = self.wb.get_sheet_by_name("move")

    def get_id(self):
        l=[]
        for i in range(1,self.sh_id.max_row+1):
            id = self.sh_id.cell(row=i,column=1).value
            l.append(id)
        return l

    # def input_excel(self,v_id,v_name,v_jj,v_dy,zhuyan_name,zhuyan_id,time):
    #     mx = self.sh_move.max_row
    #     self.sh_move.cell(row=mx + 1, column=1).value = v_id
    #     self.sh_move.cell(row=mx + 1, column=2).value = v_name
    #     self.sh_move.cell(row=mx + 1, column=3).value = v_jj
    #     self.sh_move.cell(row=mx + 1, column=4).value = v_dy
    #     self.sh_move.cell(row=mx + 1, column=5).value = zhuyan_name
    #     self.sh_move.cell(row=mx + 1, column=6).value = zhuyan_id
    #     self.sh_move.cell(row=mx + 1, column=7).value = time
    #     self.wb.save(self.excelpath)